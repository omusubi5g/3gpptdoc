import os
from ftplib import FTP
import openpyxl
import zipfile
import re

def sanitize_folder_name(folder_name):
    folder_name = re.sub(r'[\\/*?:"<>|]', '', folder_name)
    folder_name = folder_name.replace('\n', ' ').replace(',', ' ')
    folder_name = re.sub(r'\s+', ' ', folder_name)
    folder_name = folder_name.strip()
    return folder_name if folder_name else 'Unknown'

def list_directories(ftp):
    directories = []
    ftp.dir(lambda x: directories.append(x.split()[-1]))
    return [d for d in directories if '.' not in d]

def download_ftp_folder(ftp_host, remote_dir, local_dir):
    ftp = FTP(ftp_host)
    ftp.login()

    current_dir = "/"
    
    while True:
        print(f"\n現在のディレクトリ: {current_dir}")
        directories = list_directories(ftp)
        
        print("利用可能なディレクトリ:")
        for i, dir_name in enumerate(directories, 1):
            print(f"{i}. {dir_name}")
        print("0. このディレクトリをダウンロード")
        print("-1. 親ディレクトリに戻る")
        
        choice = input("選択してください (数字を入力): ")
        
        if choice == "0":
            remote_dir = current_dir
            break
        elif choice == "-1":
            current_dir = os.path.dirname(current_dir)
            ftp.cwd(current_dir)
        elif choice.isdigit() and 1 <= int(choice) <= len(directories):
            selected_dir = directories[int(choice) - 1]
            current_dir = os.path.join(current_dir, selected_dir)
            ftp.cwd(current_dir)
        else:
            print("無効な選択です。もう一度試してください。")

    if not os.path.exists(local_dir):
        os.makedirs(local_dir)

    ftp.cwd(remote_dir)

    files = ftp.nlst()

    excel_file = next((f for f in files if f.startswith('TDoc_List') and f.endswith('.xlsx')), None)

    if excel_file:
        local_excel_file = os.path.join(local_dir, excel_file)
        with open(local_excel_file, 'wb') as f:
            ftp.retrbinary(f"RETR {excel_file}", f.write)
        print(f"Excelファイルをダウンロードしました: {excel_file}")

        wb = openpyxl.load_workbook(local_excel_file)
        sheet = wb.active

        Type_folders = {}
        tdoc_Type_map = {}
        
        # ヘッダー行を探す
        header_row = None
        for row in sheet.iter_rows(values_only=True):
            if 'TDoc' in row and 'Type' in row:
                header_row = row
                break
        
        if header_row:
            tdoc_index = header_row.index('TDoc')
            Type_index = header_row.index('Type')
            
            for row in sheet.iter_rows(min_row=sheet.min_row + 1, values_only=True):
                if len(row) > max(tdoc_index, Type_index):
                    tdoc = row[tdoc_index]
                    Type = row[Type_index]
                    if Type and tdoc:
                        sanitized_Type = sanitize_folder_name(str(Type))
                        Type_folder = os.path.join(local_dir, sanitized_Type)
                        if not os.path.exists(Type_folder):
                            os.makedirs(Type_folder)
                        Type_folders[sanitized_Type] = Type_folder
                        tdoc_Type_map[tdoc] = sanitized_Type
        else:
            print("警告: ExcelファイルにTDocまたはType列が見つかりません。")

        for file in files:
            if file.endswith('.zip'):
                local_file = os.path.join(local_dir, file)
                with open(local_file, 'wb') as f:
                    ftp.retrbinary(f"RETR {file}", f.write)
                print(f"ダウンロード完了: {file}")

                file_name_without_ext = os.path.splitext(file)[0]
                if file_name_without_ext in tdoc_Type_map:
                    Type_val = tdoc_Type_map[file_name_without_ext]
                    Type_folder = Type_folders[Type_val]
                    os.rename(local_file, os.path.join(Type_folder, file))
                    
                    with zipfile.ZipFile(os.path.join(Type_folder, file), 'r') as zip_ref:
                        zip_ref.extractall(Type_folder)
                    print(f"{file} を {Type_val} フォルダに振り分け、解凍しました。")
                else:
                    print(f"警告: {file} に対応するTDocが見つかりません。")

    else:
        print("警告: TDoc_Listから始まるExcelファイルが見つかりません。")

    ftp.quit()

# スクリプトの使用例
if __name__ == "__main__":
    ftp_host = "ftp.3gpp.org"
    local_dir = "./downloaded_files"

    download_ftp_folder(ftp_host, "", local_dir)