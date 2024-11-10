import os
import hashlib  # 追加
from ftplib import FTP, error_perm, error_temp
import openpyxl
import zipfile
import re
import time
from socket import timeout as SocketTimeout
def sanitize_folder_name(folder_name, max_length=50):
    # 無効な文字を削除
    folder_name = re.sub(r'[\\/*?:"<>|]', '', folder_name)
    folder_name = folder_name.replace('\n', ' ').replace(',', ' ')
    folder_name = re.sub(r'\s+', ' ', folder_name)
    folder_name = folder_name.strip()

    # 長すぎる名前を短縮
    if len(folder_name) > max_length:
        hash_object = hashlib.md5(folder_name.encode())
        folder_name = folder_name[:max_length-8] + hash_object.hexdigest()[:8]

    return folder_name if folder_name else 'Unknown'

def list_directories(ftp):
    directories = []
    ftp.dir(lambda x: directories.append(x.split()[-1]))
    return [d for d in directories if '.' not in d]
def ftp_connect(ftp_host, max_attempts=3, delay=5):
    for attempt in range(max_attempts):
        try:
            ftp = FTP(ftp_host)
            ftp.login()
            ftp.set_pasv(True)
            ftp.encoding = 'utf-8'
            return ftp
        except (error_perm, error_temp, SocketTimeout) as e:
            print(f"接続エラー（試行 {attempt + 1}/{max_attempts}）: {str(e)}")
            if attempt < max_attempts - 1:
                print(f"{delay}秒後に再試行します...")
                time.sleep(delay)
            else:
                raise Exception("FTP接続に失敗しました。")
def refresh_connection(ftp, ftp_host):
    try:
        ftp.voidcmd("NOOP")
    except:
        print("接続が切断されました。再接続を試みます...")
        ftp.close()
        return ftp_connect(ftp_host)
    return ftp
def download_file(ftp, remote_file, local_file, ftp_host, max_attempts=3, delay=5):
    for attempt in range(max_attempts):
        try:
            ftp = refresh_connection(ftp, ftp_host)
            with open(local_file, 'wb') as f:
                ftp.retrbinary(f"RETR {remote_file}", f.write)
            return ftp, True
        except (error_perm, error_temp, SocketTimeout, ConnectionResetError) as e:
            print(f"ダウンロードエラー（試行 {attempt + 1}/{max_attempts}）: {str(e)}")
            if attempt < max_attempts - 1:
                print(f"{delay}秒後に再試行します...")
                time.sleep(delay)
            else:
                print(f"ファイル {remote_file} のダウンロードに失敗しました。")
                return ftp, False
def download_ftp_folder(ftp_host, remote_dir, local_dir):
    ftp = ftp_connect(ftp_host)

    current_dir = "/"
    
    while True:
        ftp = refresh_connection(ftp, ftp_host)
        print(f"\n現在のディレクトリ: {current_dir}")
        directories = list_directories(ftp)
        
        print("利用可能なディレクトリ:")
        for i, dir_name in enumerate(directories, 1):
            print(f"{i}. {dir_name}")
        print("0. このディレクトリをダウンロード")
        print("-1. 親ディレクトリに戻る")
        
        choice = input("選択してください (数字を入力): ")
        
        if choice == "0":
            remote_dir = current_dir
            break
        elif choice == "-1":
            current_dir = os.path.dirname(current_dir)
            ftp.cwd(current_dir)
        elif choice.isdigit() and 1 <= int(choice) <= len(directories):
            selected_dir = directories[int(choice) - 1]
            current_dir = os.path.join(current_dir, selected_dir)
            ftp.cwd(current_dir)
        else:
            print("無効な選択です。もう一度試してください。")

    if not os.path.exists(local_dir):
        os.makedirs(local_dir)

    ftp.cwd(remote_dir)

    ftp = refresh_connection(ftp, ftp_host)
    files = ftp.nlst()

    excel_file = next((f for f in files if f.startswith('TDoc_List') and f.endswith('.xlsx')), None)

    if excel_file:
        local_excel_file = os.path.join(local_dir, excel_file)
        ftp, success = download_file(ftp, excel_file, local_excel_file, ftp_host)
        if success:
            print(f"Excelファイルをダウンロードしました: {excel_file}")

            wb = openpyxl.load_workbook(local_excel_file)
            sheet = wb.active

            type_folders = {}
            tdoc_type_map = {}

            header_row = None
            for row in sheet.iter_rows(values_only=True):
                if 'TDoc' in row and 'Agenda item' in row:
                    header_row = row
                    break

            if header_row:
                tdoc_index = header_row.index('TDoc')
                type_index = header_row.index('Agenda item')

                for row in sheet.iter_rows(min_row=sheet.min_row + 1, values_only=True):
                    if len(row) > max(tdoc_index, type_index):
                        tdoc = row[tdoc_index]
                        type_val = row[type_index]
                        if type_val and tdoc:
                            sanitized_type = sanitize_folder_name(str(type_val))
                            type_folder = os.path.join(local_dir, sanitized_type)
                            try:
                                if not os.path.exists(type_folder):
                                    os.makedirs(type_folder)
                                type_folders[sanitized_type] = type_folder
                                tdoc_type_map[tdoc] = sanitized_type
                            except OSError as e:
                                print(f"フォルダの作成に失敗しました: {type_folder}")
                                print(f"エラー: {str(e)}")

                for file in files:
                    if file.endswith('.zip'):
                        local_file = os.path.join(local_dir, file)
                        ftp, success = download_file(ftp, file, local_file, ftp_host)
                        if success:
                            print(f"ダウンロード完了: {file}")

                            file_name_without_ext = os.path.splitext(file)[0]
                            if file_name_without_ext in tdoc_type_map:
                                type_val = tdoc_type_map[file_name_without_ext]
                                type_folder = type_folders[type_val]
                                try:
                                    os.rename(local_file, os.path.join(type_folder, file))
                                    
                                    with zipfile.ZipFile(os.path.join(type_folder, file), 'r') as zip_ref:
                                        zip_ref.extractall(type_folder)
                                    print(f"{file} を {type_val} フォルダに振り分け、解凍しました。")
                                except OSError as e:
                                    print(f"ファイルの移動または解凍に失敗しました: {file}")
                                    print(f"エラー: {str(e)}")
                            else:
                                print(f"警告: {file} に対応するTDocが見つかりません。")
            else:
                print("警告: ExcelファイルにTDocまたはType列が見つかりません。")
        else:
            print("Excelファイルのダウンロードに失敗しました。")
    else:
        print("警告: TDoc_Listから始まるExcelファイルが見つかりません。")

    ftp.quit()
if __name__ == "__main__":
    ftp_host = "ftp.3gpp.org"
    local_dir = "./downloaded_files"
    download_ftp_folder(ftp_host, "", local_dir)
